"""Produce an xlsx report from the `walkforward_bets` DuckDB table.

Writes one workbook with:
- Summary: run metadata, grand totals (bets, wins, losses, stake, payout,
  ROI, hit rate), all computed as Excel formulas against the bet log.
- Monthly: per-month totals (bets, wins, losses, net payout, ROI).
- Yearly: per-year totals.
- By Edge Bucket: rollup by 1pp edge buckets (0.03-0.04, 0.04-0.05, ...).
- All Bets: full bet log, one row per flagged bet.

After writing, formulas are recalculated via `scripts/recalc.py` so the
summary tabs show live values when opened.
"""

from __future__ import annotations

import logging
import subprocess
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from mlpm.config.settings import settings
from mlpm.storage.duckdb import connect_read_only, query_dataframe

logger = logging.getLogger(__name__)

BOLD = Font(bold=True)
HEADER_FILL = PatternFill("solid", start_color="1F3864", end_color="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _load_bets(run_id: str | None, db_path: Path | None = None) -> pd.DataFrame:
    conn = connect_read_only(db_path or settings().duckdb_path)
    try:
        if run_id:
            return query_dataframe(
                conn,
                """
                SELECT * FROM walkforward_bets
                WHERE run_id = ?
                ORDER BY game_date, game_id
                """,
                (run_id,),
            )
        return query_dataframe(
            conn,
            """
            SELECT * FROM walkforward_bets
            WHERE run_id = (
                SELECT run_id FROM walkforward_bets
                ORDER BY scored_at DESC
                LIMIT 1
            )
            ORDER BY game_date, game_id
            """,
        )
    finally:
        conn.close()


def _style_header(sheet, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = sheet.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def _autosize(sheet, columns: list[str]) -> None:
    for idx, column in enumerate(columns, start=1):
        header_width = len(str(column))
        sheet.column_dimensions[get_column_letter(idx)].width = max(12, header_width + 2)


def _write_summary(wb: Workbook, bets: pd.DataFrame, run_id: str, bets_sheet_name: str) -> None:
    sheet = wb.create_sheet("Summary", index=0)
    sheet["A1"] = "Walk-Forward Backtest Report"
    sheet["A1"].font = Font(bold=True, size=14)
    model_names = sorted(bets["model_name"].dropna().astype(str).unique().tolist()) if not bets.empty else []
    model_display = ", ".join(model_names) if model_names else "-"

    meta_rows = [
        ("Run ID", run_id),
        ("Model", model_display),
        ("Model count", len(model_names)),
        ("Bets", len(bets)),
        ("Earliest bet", str(bets["game_date"].min()) if not bets.empty else "-"),
        ("Latest bet", str(bets["game_date"].max()) if not bets.empty else "-"),
        ("Report generated", datetime.utcnow().isoformat(timespec="seconds") + "Z"),
    ]
    row = 3
    for label, value in meta_rows:
        sheet.cell(row=row, column=1, value=label).font = BOLD
        sheet.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    # Aggregates — formulas so the workbook stays live if rows are edited.
    n = len(bets)
    if n == 0:
        sheet.cell(row=row, column=1, value="No bets in this run.").font = BOLD
        return

    bet_log_range_won = f"'{bets_sheet_name}'!N2:N{n + 1}"      # won_bet column
    bet_log_range_stake = f"'{bets_sheet_name}'!M2:M{n + 1}"    # stake column
    bet_log_range_payout = f"'{bets_sheet_name}'!O2:O{n + 1}"   # payout column
    bet_log_range_edge = f"'{bets_sheet_name}'!K2:K{n + 1}"     # edge_pct column

    sheet.cell(row=row, column=1, value="Grand Totals").font = BOLD
    row += 1
    aggregate_rows = [
        ("Total bets", f"=COUNTA({bet_log_range_won})", "0"),
        ("Wins", f"=COUNTIF({bet_log_range_won},TRUE)", "0"),
        ("Losses", f"=COUNTIF({bet_log_range_won},FALSE)", "0"),
        ("Hit rate", f"=COUNTIF({bet_log_range_won},TRUE)/COUNTA({bet_log_range_won})", "0.00%"),
        ("Total stake", f"=SUM({bet_log_range_stake})", "0.00"),
        ("Total payout (net)", f"=SUM({bet_log_range_payout})", "0.00"),
        ("ROI", f"=SUM({bet_log_range_payout})/SUM({bet_log_range_stake})", "0.00%"),
        ("Avg edge (pp)", f"=AVERAGE({bet_log_range_edge})", "0.0000"),
        ("Break-even hit rate at avg price",
            f"=1/AVERAGEIF({bet_log_range_won},FALSE,'{bets_sheet_name}'!L2:L{n + 1})"
            f" + 0",  # placeholder; overridden below
         "0.00%"),
    ]
    # Overwrite the "break-even" formula to something more meaningful — the
    # implied break-even rate at the average decimal odds across all bets.
    aggregate_rows[-1] = (
        "Break-even hit rate (avg odds)",
        f"=1/AVERAGE('{bets_sheet_name}'!L2:L{n + 1})",
        "0.00%",
    )
    for label, formula, number_format in aggregate_rows:
        sheet.cell(row=row, column=1, value=label).font = BOLD
        cell = sheet.cell(row=row, column=2, value=formula)
        cell.number_format = number_format
        row += 1

    sheet.column_dimensions["A"].width = 32
    sheet.column_dimensions["B"].width = 24


def _write_bet_log(wb: Workbook, bets: pd.DataFrame, sheet_name: str) -> None:
    sheet = wb.create_sheet(sheet_name)
    columns = [
        "run_id", "model_name", "game_id", "game_date",
        "home_team", "away_team", "picked_team", "is_home_pick",
        "model_prob", "market_prob", "edge_pct", "decimal_odds",
        "stake", "won_bet", "payout", "winner_team",
        "train_rows", "train_start_date", "train_end_date", "scored_at",
    ]
    # Write header
    for c, col in enumerate(columns, start=1):
        sheet.cell(row=1, column=c, value=col)
    _style_header(sheet, 1, len(columns))
    # Write data
    for r, record in enumerate(bets[columns].to_dict(orient="records"), start=2):
        for c, col in enumerate(columns, start=1):
            value = record[col]
            if isinstance(value, pd.Timestamp):
                value = value.to_pydatetime()
            sheet.cell(row=r, column=c, value=value)
    _autosize(sheet, columns)
    sheet.freeze_panes = "A2"


def _write_grouped_summary(
    wb: Workbook,
    bets: pd.DataFrame,
    *,
    sheet_name: str,
    group_label: str,
    grouper: pd.Series,
) -> None:
    sheet = wb.create_sheet(sheet_name)
    # Aggregate in pandas for layout; then write values (these are reports not
    # interactive models, and the cross-sheet formula approach gets gnarly).
    grp = bets.assign(_g=grouper).groupby("_g", dropna=False)
    table = pd.DataFrame(
        {
            "bets": grp.size(),
            "wins": grp["won_bet"].sum().astype(int),
            "losses": (grp.size() - grp["won_bet"].sum().astype(int)).astype(int),
            "stake": grp["stake"].sum(),
            "payout": grp["payout"].sum(),
            "avg_edge_pct": grp["edge_pct"].mean(),
            "avg_decimal_odds": grp["decimal_odds"].mean(),
        }
    ).reset_index().rename(columns={"_g": group_label})
    # Derived with formulas so users can spot-check in-sheet.
    table["hit_rate"] = table["wins"] / table["bets"]
    table["roi"] = table["payout"] / table["stake"]

    columns = list(table.columns)
    for c, col in enumerate(columns, start=1):
        sheet.cell(row=1, column=c, value=col)
    _style_header(sheet, 1, len(columns))
    for r, record in enumerate(table.to_dict(orient="records"), start=2):
        for c, col in enumerate(columns, start=1):
            value = record[col]
            if isinstance(value, pd.Timestamp):
                value = value.to_pydatetime()
            cell = sheet.cell(row=r, column=c, value=value)
            if col in {"hit_rate", "roi"}:
                cell.number_format = "0.00%"
            elif col in {"stake", "payout"}:
                cell.number_format = "0.00"
            elif col in {"avg_edge_pct", "avg_decimal_odds"}:
                cell.number_format = "0.0000"
    _autosize(sheet, columns)
    sheet.freeze_panes = "A2"


def generate_walkforward_report(
    output_path: Path | str,
    *,
    run_id: str | None = None,
    db_path: Path | None = None,
    recalc: bool = True,
) -> dict:
    """Write the walkforward report xlsx. Returns summary dict with totals."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    bets = _load_bets(run_id, db_path=db_path)
    if bets.empty:
        logger.warning("No bets found for run_id=%s; writing empty workbook.", run_id)

    wb = Workbook()
    # Remove the default sheet so the first sheet is "Summary".
    wb.remove(wb.active)
    bets_sheet_name = "All Bets"
    resolved_run_id = str(bets["run_id"].iloc[0]) if not bets.empty else (run_id or "")

    _write_bet_log(wb, bets, bets_sheet_name)
    _write_summary(wb, bets, run_id=resolved_run_id, bets_sheet_name=bets_sheet_name)
    if not bets.empty:
        _write_grouped_summary(
            wb,
            bets,
            sheet_name="By Model",
            group_label="model_name",
            grouper=bets["model_name"].astype(str),
        )
        _write_grouped_summary(
            wb, bets,
            sheet_name="Monthly",
            group_label="month",
            grouper=pd.to_datetime(bets["game_date"]).dt.to_period("M").astype(str),
        )
        _write_grouped_summary(
            wb, bets,
            sheet_name="Yearly",
            group_label="year",
            grouper=pd.to_datetime(bets["game_date"]).dt.year,
        )
        # Edge buckets start at 0pp so reports remain accurate even when users
        # run the backtest with a lower threshold than the default 3pp.
        bins = [0.0, 0.04, 0.05, 0.06, 0.08, 0.10, 1.0]
        labels = ["0-4pp", "4-5pp", "5-6pp", "6-8pp", "8-10pp", "10pp+"]
        edge_bucket = pd.cut(bets["edge_pct"], bins=bins, labels=labels, include_lowest=True)
        _write_grouped_summary(
            wb, bets,
            sheet_name="By Edge Bucket",
            group_label="edge_bucket",
            grouper=edge_bucket.astype(str),
        )

    wb.save(output_path)

    result = {
        "status": "ok",
        "path": str(output_path),
        "bets": int(len(bets)),
        "run_id": resolved_run_id,
    }

    if recalc and not bets.empty:
        # Try a few common locations. If none found, skip silently — Excel
        # auto-recalculates on open, so this step is only needed when the
        # user wants the rendered xlsx to already contain computed values.
        candidate_scripts = [
            Path(__file__).resolve().parents[3] / "scripts" / "recalc.py",
            Path.home() / ".claude" / "skills" / "xlsx" / "scripts" / "recalc.py",
        ]
        recalc_script = next((p for p in candidate_scripts if p.exists()), None)
        if recalc_script is not None:
            try:
                subprocess.run(
                    ["python", str(recalc_script), str(output_path), "60"],
                    check=True,
                    capture_output=True,
                    text=True,
                    timeout=120,
                )
                result["recalculated"] = True
            except (subprocess.CalledProcessError, subprocess.TimeoutExpired) as exc:
                logger.info("recalc step skipped (non-fatal): %s", exc)
                result["recalculated"] = False
        else:
            result["recalculated"] = False

    return result
